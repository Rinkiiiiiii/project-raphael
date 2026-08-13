"""
main_script.py
---------------
Logic layer for Raphael's "excel_graph_creation" sub-program.

Responsible for:
    - Defining the data model for the three supported graph types
      (Behavior, Replacement, Caregiver's Goal).
    - Validating user-supplied data, including the client name used to
      name the workbook file (e.g. "John_Smith_raphael_behavioral_data.xlsx").
    - Maintaining ONE shared workbook per client, with one sheet per
      graph type. Each sheet holds many tables, stacked in fixed-height
      "slots" (30 rows apart) so tables never collide.
    - Locating the right slot for a given category name (creating a
      new one, or reusing/overwriting an existing one for the same
      category), building the data table + chart in that slot.
    - Saving the workbook and opening it in the user's default
      spreadsheet application.

This module has NO knowledge of any UI toolkit. main_script_ui.py
(or any other front-end) is expected to build a GraphRequest, call
generate_excel_graph(), and optionally call open_file() on the
result.
"""

from __future__ import annotations

import os
import re
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import date
from enum import Enum
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.plotarea import DataTable
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.exceptions import CellCoordinatesException
from openpyxl.worksheet.worksheet import Worksheet


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #

class GraphType(Enum):
    """The three graph types Raphael currently supports."""

    BEHAVIOR = "Behavior"
    REPLACEMENT = "Replacement"
    CAREGIVER_GOAL = "Caregiver's Goal"

    @property
    def value_label(self) -> str:
        """Label for the numeric column/axis for this graph type."""
        return "Frequency" if self is GraphType.BEHAVIOR else "Percentage"

    @property
    def category_label(self) -> str:
        """Label describing what the user names for this graph type."""
        return {
            GraphType.BEHAVIOR: "Type of Behavior",
            GraphType.REPLACEMENT: "Type of Replacement",
            GraphType.CAREGIVER_GOAL: "Caregiver's Goal",
        }[self]

    @property
    def is_percentage(self) -> bool:
        return self is not GraphType.BEHAVIOR

    @property
    def sheet_name(self) -> str:
        """Fixed, stable tab name for this graph type within the shared workbook."""
        return self.value


@dataclass
class DataPoint:
    """A single (date, value) observation."""

    entry_date: date
    value: float  # raw frequency count, OR a percentage expressed 0-100


@dataclass
class GraphRequest:
    """Everything needed to build (or replace) one category's table + graph."""

    graph_type: GraphType
    category_name: str                       # e.g. "Elopement", "Requests a Break"
    client_first_name: str = ""              # used to name the workbook file
    client_last_name: str = ""               # used to name the workbook file
    data_points: List[DataPoint] = field(default_factory=list)
    this_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(this_dir, "..", "client_behavioral_data_files")        # folder holding the shared workbook
    workbook_path: Optional[str] = None       # override the default shared-file path


@dataclass
class GraphResult:
    """What generate_excel_graph() hands back so a UI can report what happened."""

    workbook_path: str
    sheet_name: str
    anchor_row: int          # the row the table's header landed on
    was_replacement: bool    # True if this overwrote an existing table for the category


class GraphRequestError(ValueError):
    """Raised when a GraphRequest fails validation."""


# --------------------------------------------------------------------------- #
# Layout constants
# --------------------------------------------------------------------------- #

WORKBOOK_FILENAME_SUFFIX = "Raphael_Behavioral_Data.xlsx"

ROWS_PER_SLOT = 30           # every table starts exactly this many rows after the last
MAX_DATA_POINTS_PER_TABLE = ROWS_PER_SLOT - 3   # header + data rows, with a buffer left over
_MAX_SLOT_SEARCH_ITERATIONS = 10_000             # safety cap; a healthy sheet never gets close

_DATA_COL = 1        # column A: Date
_VALUE_COL = 2        # column B: category value (frequency or percentage)
_CHART_ANCHOR_COL_OFFSET = 3  # chart is anchored 3 columns right of the Date column (column D)

# How many columns wide a slot's formatting-clear should reach when a table
# is overwritten -- wide enough to also wipe any leftover formatting from
# an older version of this file (e.g. the beige gap-band this program used
# to draw), even though nothing new is ever painted out that far anymore.
_SLOT_FILL_LAST_COL = 14  # column N

CHART_HEIGHT_CM = 3.0 * 2.54    # 3 in
CHART_WIDTH_CM = 6.5 * 2.54     # 6.5 in


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #

def validate_request(request: GraphRequest) -> None:
    """
    Raises GraphRequestError with a human-readable message if the
    request cannot be turned into a graph. Called automatically by
    generate_excel_graph, but exposed so the UI can validate early
    and show inline errors before attempting to generate anything.
    """
    if request.category_name is None or not request.category_name.strip():
        raise GraphRequestError(f"{request.graph_type.category_label} cannot be empty.")

    # The client's name is only required to name the workbook file -- if the
    # caller supplied an explicit workbook_path, that requirement is moot.
    if not request.workbook_path:
        if not request.client_first_name or not request.client_first_name.strip():
            raise GraphRequestError("Client first name is required.")
        if not request.client_last_name or not request.client_last_name.strip():
            raise GraphRequestError("Client last name is required.")

    if not request.data_points:
        raise GraphRequestError("At least one data point (date + value) is required.")

    if len(request.data_points) < 2:
        raise GraphRequestError(
            "At least two data points are required to plot a meaningful trend line."
        )

    if len(request.data_points) > MAX_DATA_POINTS_PER_TABLE:
        raise GraphRequestError(
            f"A single table supports at most {MAX_DATA_POINTS_PER_TABLE} data points "
            f"(so tables stay {ROWS_PER_SLOT} rows apart and never collide). "
            f"You supplied {len(request.data_points)}."
        )

    for point in request.data_points:
        if not isinstance(point.entry_date, date):
            raise GraphRequestError(f"Invalid date: {point.entry_date!r}")

        if request.graph_type.is_percentage:
            if not (0 <= point.value <= 100):
                raise GraphRequestError(
                    f"Percentage value {point.value} on {point.entry_date} "
                    "must be between 0 and 100."
                )
        else:
            if point.value < 0:
                raise GraphRequestError(
                    f"Frequency value {point.value} on {point.entry_date} cannot be negative."
                )


# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Arial", size=11)

_BAND_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")   # light blue
_NO_FILL = PatternFill(fill_type=None)


# --------------------------------------------------------------------------- #
# Workbook / sheet setup
# --------------------------------------------------------------------------- #

def _sanitize_name_part(name: str) -> str:
    """Strips filesystem-unsafe characters and collapses whitespace to underscores."""
    cleaned = re.sub(r'[\\/*?:"<>|]', "", name).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned


def build_workbook_filename(client_first_name: str, client_last_name: str) -> str:
    """
    Returns the standard filename for a client's workbook, e.g.
    "John_Smith_raphael_behavioral_data.xlsx". Exposed publicly so a UI
    can preview the filename before generating anything.
    """
    first = _sanitize_name_part(client_first_name)
    last = _sanitize_name_part(client_last_name)
    return f"{first}_{last}_{WORKBOOK_FILENAME_SUFFIX}"


def _resolve_workbook_path(request: GraphRequest) -> str:
    if request.workbook_path:
        return os.path.abspath(request.workbook_path)
    filename = build_workbook_filename(request.client_first_name, request.client_last_name)
    return os.path.abspath(os.path.join(request.output_dir, filename))


def _load_or_create_workbook(workbook_path: str) -> Workbook:
    if os.path.exists(workbook_path):
        return load_workbook(workbook_path)

    wb = Workbook()
    # Rename the default sheet to the first graph type instead of leaving
    # a stray "Sheet" tab, then add the other two.
    all_types = list(GraphType)
    wb.active.title = all_types[0].sheet_name
    for graph_type in all_types[1:]:
        wb.create_sheet(graph_type.sheet_name)
    return wb


def _get_or_create_sheet(wb: Workbook, graph_type: GraphType) -> Worksheet:
    if graph_type.sheet_name not in wb.sheetnames:
        wb.create_sheet(graph_type.sheet_name)
    return wb[graph_type.sheet_name]


# --------------------------------------------------------------------------- #
# Slot search — the core "find matching table, else find free space" logic
# --------------------------------------------------------------------------- #

def _find_or_create_slot(ws: Worksheet, category_name: str) -> tuple[int, bool]:
    """
    Scans column B at rows 1, 31, 61, ... (each table's header row) looking
    for an existing table whose category name matches. Returns
    (anchor_row, is_new_slot).

    The exit case is checked FIRST, before any text comparison: an empty
    header cell means we've reached the end of the sheet's existing tables
    (tables are always written contiguously, slot by slot, with no gaps),
    so that row is free to use and the search stops there. This both
    prevents comparing against an empty cell and guarantees termination
    without needing to know how many tables already exist.
    """
    normalized_target = category_name.strip().lower()
    row = 1
    iterations = 0

    while True:
        iterations += 1
        if iterations > _MAX_SLOT_SEARCH_ITERATIONS:
            raise RuntimeError(
                "Slot search exceeded its safety limit without finding free space. "
                "The sheet may have a gap or corrupted layout — please check it manually."
            )

        header_cell = ws.cell(row=row, column=_VALUE_COL)

        # Exit case: an empty header cell means this slot (and everything
        # after it) is unused — stop here rather than comparing text.
        if header_cell.value is None or str(header_cell.value).strip() == "":
            return row, True

        if str(header_cell.value).strip().lower() == normalized_target:
            return row, False

        row += ROWS_PER_SLOT


def _chart_anchor_row(chart: LineChart) -> Optional[int]:
    """
    Returns the 1-indexed row a chart is anchored to, or None if unknown.

    A chart we just added in this process still has a plain string anchor
    (e.g. "D1"). A chart that came from load_workbook() on a file saved
    earlier has already been parsed into a OneCellAnchor/TwoCellAnchor
    object, whose _from.row is 0-indexed. Both cases have to be handled,
    or a reloaded workbook's existing charts become invisible to the
    slot-clearing logic and get left behind as orphans.
    """
    anchor = chart.anchor
    if isinstance(anchor, str):
        try:
            _, row = coordinate_from_string(anchor)
            return row
        except (ValueError, CellCoordinatesException):
            return None
    if hasattr(anchor, "_from") and anchor._from is not None:
        return anchor._from.row + 1
    return None


def _clear_slot(ws: Worksheet, anchor_row: int) -> None:
    """
    Wipes everything belonging to a previously-generated table in this
    slot: cell values/formatting across the full slot height and the
    wide clear-column range (see _SLOT_FILL_LAST_COL), and any chart
    anchored within the slot's row range.
    """
    for r in range(anchor_row, anchor_row + ROWS_PER_SLOT):
        for c in range(_DATA_COL, _SLOT_FILL_LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.comment = None
            cell.number_format = "General"
            cell.fill = _NO_FILL

    slot_end = anchor_row + ROWS_PER_SLOT
    ws._charts = [
        c for c in ws._charts
        if not (_chart_anchor_row(c) is not None and anchor_row <= _chart_anchor_row(c) < slot_end)
    ]


# --------------------------------------------------------------------------- #
# Table + chart construction within a slot
# --------------------------------------------------------------------------- #

def _write_table(ws: Worksheet, request: GraphRequest, anchor_row: int) -> tuple[int, int]:
    """
    Writes the header + sorted data rows starting at anchor_row.
    Returns (first_data_row, last_data_row).
    """
    graph_type = request.graph_type

    header_cell_a = ws.cell(row=anchor_row, column=_DATA_COL, value="Date")
    header_cell_b = ws.cell(row=anchor_row, column=_VALUE_COL, value=request.category_name.strip())
    for cell in (header_cell_a, header_cell_b):
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    sorted_points = sorted(request.data_points, key=lambda p: p.entry_date)
    first_data_row = anchor_row + 1

    for i, point in enumerate(sorted_points):
        r = first_data_row + i
        row_fill = _BAND_FILL if i % 2 == 0 else _NO_FILL

        date_cell = ws.cell(row=r, column=_DATA_COL, value=point.entry_date)
        date_cell.number_format = "mm/dd/yyyy"
        date_cell.font = _BODY_FONT
        date_cell.fill = row_fill

        if graph_type.is_percentage:
            value_cell = ws.cell(row=r, column=_VALUE_COL, value=point.value / 100.0)
            value_cell.number_format = "0.0%"
        else:
            value_cell = ws.cell(row=r, column=_VALUE_COL, value=point.value)
            value_cell.number_format = "0"
        value_cell.font = _BODY_FONT
        value_cell.fill = row_fill

    ws.column_dimensions[get_column_letter(_DATA_COL)].width = 14
    ws.column_dimensions[get_column_letter(_VALUE_COL)].width = max(18, len(request.category_name) + 4)

    last_data_row = first_data_row + len(sorted_points) - 1
    return first_data_row, last_data_row


def _no_bold(title) -> None:
    """
    Strips the bold-by-default weight Excel/LibreOffice apply to chart
    and axis titles, so they render as regular text instead.
    """
    title.tx.rich.p[0].pPr.defRPr.b = False


def _add_chart(ws: Worksheet, request: GraphRequest, anchor_row: int,
                first_data_row: int, last_data_row: int) -> None:
    """Builds and inserts a line chart for this slot's table."""
    chart = LineChart()
    chart.title = request.category_name.strip()
    chart.style = 2
    chart.type = "line"
    chart.grouping = "standard"
    chart.marker = True

    # A visible chart-area border/frame is drawn by default unless
    # explicitly turned off.
    chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    _no_bold(chart.title)

    # No axis title/tick labels for dates -- the data table below the plot
    # (added further down) already shows every date, so a second copy on
    # the axis itself would just be redundant.
    chart.x_axis.axPos = "b"          # bottom -- openpyxl defaults this to "l", which is wrong
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "none"
    chart.x_axis.majorGridlines = None

    chart.y_axis.title = request.graph_type.value_label
    chart.y_axis.axPos = "l"          # left -- also defaults to "l", but needs to be explicit
    chart.y_axis.delete = False
    # Thin, light-gray gridlines instead of the heavy black default.
    chart.y_axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=9525))
    )
    _no_bold(chart.y_axis.title)

    chart.height = CHART_HEIGHT_CM   # 3 in
    chart.width = CHART_WIDTH_CM     # 6.5 in

    # Leave visible margin around the plotted line so the axis titles, tick
    # labels, and chart title all have breathing room instead of the data
    # running edge-to-edge.
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="inner",
            xMode="edge", yMode="edge",
            x=0.14, y=0.16, w=0.68, h=0.50,
        )
    )

    if request.graph_type.is_percentage:
        chart.y_axis.number_format = "0%"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1

    values = Reference(ws, min_col=_VALUE_COL, min_row=anchor_row, max_row=last_data_row)
    categories = Reference(ws, min_col=_DATA_COL, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    series = chart.series[0]
    series.marker.symbol = "circle"
    series.smooth = False

    # A data table under the plot (dates across the top, values below, a
    # legend key in the corner) replaces the need for a separate floating
    # legend or x-axis date labels.
    chart.plot_area.dTable = DataTable(
        showHorzBorder=True, showVertBorder=True, showOutline=True, showKeys=True
    )
    chart.legend = None

    anchor_col = get_column_letter(_DATA_COL + _CHART_ANCHOR_COL_OFFSET)
    ws.add_chart(chart, f"{anchor_col}{anchor_row}")


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #

def generate_excel_graph(request: GraphRequest) -> GraphResult:
    """
    Validates the request, then:
      1. Opens the shared workbook (creating it, with all three sheets,
         if it doesn't exist yet).
      2. On the sheet matching request.graph_type, searches existing
         table slots (row 1, 31, 61, ...) for one whose category name
         already matches request.category_name.
      3. If found, clears that slot's old table/chart and rewrites it
         with the new data. If not found, uses the first free slot.
      4. Saves the workbook.

    Returns a GraphResult describing where the table landed.
    Raises GraphRequestError on invalid input, OSError if the file
    cannot be written.
    """
    validate_request(request)

    workbook_path = _resolve_workbook_path(request)
    os.makedirs(os.path.dirname(workbook_path) or ".", exist_ok=True)

    wb = _load_or_create_workbook(workbook_path)
    ws = _get_or_create_sheet(wb, request.graph_type)

    anchor_row, is_new_slot = _find_or_create_slot(ws, request.category_name)
    if not is_new_slot:
        _clear_slot(ws, anchor_row)

    first_data_row, last_data_row = _write_table(ws, request, anchor_row)
    _add_chart(ws, request, anchor_row, first_data_row, last_data_row)

    wb.save(workbook_path)

    return GraphResult(
        workbook_path=workbook_path,
        sheet_name=ws.title,
        anchor_row=anchor_row,
        was_replacement=not is_new_slot,
    )


# --------------------------------------------------------------------------- #
# Opening the result for the user
# --------------------------------------------------------------------------- #

def open_file(path: str) -> None:
    """Opens the given file with the OS default application, cross-platform."""
    if sys.platform.startswith("darwin"):
        subprocess.run(["open", path], check=False)
    elif os.name == "nt":
        os.startfile(path)  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", path], check=False)