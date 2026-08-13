"""
test_main_script.py
--------------------
Test suite for Raphael's "excel_graph_creation" logic layer
(main_script.py).

Run with:
    python -m unittest test_main_script.py -v
or:
    pytest test_main_script.py -v

Two kinds of coverage live here:

  1. Unit tests — each exercise one function/method in isolation
     (validation, slot search, slot clearing, table writing, chart
     building, chart-anchor parsing, file opening). These use
     temporary directories/workbooks that are thrown away after
     each test.

  2. A full backend integration test (TestFullBackendReport) that
     drives generate_excel_graph() the way an external caller would:
     many categories, multiple rows each, across all three graph
     types, including an overwrite. Its output is written to
     tests_output/excel_graph_creation_tests.xlsx next to this file
     and is NOT deleted afterward, so it can be opened and reviewed
     by eye.
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
import unittest
from datetime import date, timedelta
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

import main_script as ms


THIS_DIR = os.path.dirname(os.path.abspath(__file__))
TESTS_OUTPUT_DIR = os.path.join(THIS_DIR, "tests_output")
FULL_BACKEND_FILENAME = "excel_graph_creation_tests.xlsx"


def make_points(count: int, start: date = date(2026, 1, 1), step_days: int = 7,
                 value_fn=lambda i: i + 1) -> list:
    return [
        ms.DataPoint(entry_date=start + timedelta(days=step_days * i), value=value_fn(i))
        for i in range(count)
    ]


# --------------------------------------------------------------------------- #
# GraphType
# --------------------------------------------------------------------------- #

class TestGraphTypeProperties(unittest.TestCase):

    def test_value_label(self):
        self.assertEqual(ms.GraphType.BEHAVIOR.value_label, "Frequency")
        self.assertEqual(ms.GraphType.REPLACEMENT.value_label, "Percentage")
        self.assertEqual(ms.GraphType.CAREGIVER_GOAL.value_label, "Percentage")

    def test_category_label(self):
        self.assertEqual(ms.GraphType.BEHAVIOR.category_label, "Type of Behavior")
        self.assertEqual(ms.GraphType.REPLACEMENT.category_label, "Type of Replacement")
        self.assertEqual(ms.GraphType.CAREGIVER_GOAL.category_label, "Caregiver's Goal")

    def test_is_percentage(self):
        self.assertFalse(ms.GraphType.BEHAVIOR.is_percentage)
        self.assertTrue(ms.GraphType.REPLACEMENT.is_percentage)
        self.assertTrue(ms.GraphType.CAREGIVER_GOAL.is_percentage)

    def test_sheet_name_matches_value(self):
        for graph_type in ms.GraphType:
            self.assertEqual(graph_type.sheet_name, graph_type.value)


# --------------------------------------------------------------------------- #
# validate_request
# --------------------------------------------------------------------------- #

class TestValidateRequest(unittest.TestCase):

    def _behavior_request(self, **overrides) -> ms.GraphRequest:
        defaults = dict(
            graph_type=ms.GraphType.BEHAVIOR,
            category_name="Tantrum",
            client_first_name="John",
            client_last_name="Smith",
            data_points=make_points(3),
        )
        defaults.update(overrides)
        return ms.GraphRequest(**defaults)

    def test_valid_behavior_request_passes(self):
        ms.validate_request(self._behavior_request())  # should not raise

    def test_valid_percentage_request_passes(self):
        req = ms.GraphRequest(
            graph_type=ms.GraphType.REPLACEMENT,
            category_name="Uses Words",
            client_first_name="John",
            client_last_name="Smith",
            data_points=make_points(3, value_fn=lambda i: 50 + i),
        )
        ms.validate_request(req)  # should not raise

    def test_empty_category_raises(self):
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(category_name="   "))

    def test_missing_client_first_name_raises(self):
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(client_first_name="  "))

    def test_missing_client_last_name_raises(self):
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(client_last_name=""))

    def test_client_name_not_required_when_workbook_path_given(self):
        req = self._behavior_request(
            client_first_name="", client_last_name="", workbook_path="/tmp/explicit.xlsx"
        )
        ms.validate_request(req)  # should not raise

    def test_no_data_points_raises(self):
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(data_points=[]))

    def test_single_data_point_raises(self):
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(data_points=make_points(1)))

    def test_too_many_data_points_raises(self):
        too_many = make_points(ms.MAX_DATA_POINTS_PER_TABLE + 1)
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(data_points=too_many))

    def test_max_data_points_exactly_at_limit_passes(self):
        exactly_max = make_points(ms.MAX_DATA_POINTS_PER_TABLE)
        ms.validate_request(self._behavior_request(data_points=exactly_max))  # should not raise

    def test_negative_frequency_raises(self):
        points = [ms.DataPoint(date(2026, 1, 1), -1), ms.DataPoint(date(2026, 1, 2), 2)]
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(data_points=points))

    def test_percentage_above_100_raises(self):
        req = ms.GraphRequest(
            graph_type=ms.GraphType.CAREGIVER_GOAL,
            category_name="Goal",
            client_first_name="John",
            client_last_name="Smith",
            data_points=[ms.DataPoint(date(2026, 1, 1), 101), ms.DataPoint(date(2026, 1, 2), 50)],
        )
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(req)

    def test_percentage_below_0_raises(self):
        req = ms.GraphRequest(
            graph_type=ms.GraphType.CAREGIVER_GOAL,
            category_name="Goal",
            client_first_name="John",
            client_last_name="Smith",
            data_points=[ms.DataPoint(date(2026, 1, 1), -5), ms.DataPoint(date(2026, 1, 2), 50)],
        )
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(req)

    def test_invalid_date_type_raises(self):
        points = [ms.DataPoint("2026-01-01", 5), ms.DataPoint(date(2026, 1, 2), 6)]
        with self.assertRaises(ms.GraphRequestError):
            ms.validate_request(self._behavior_request(data_points=points))

# --------------------------------------------------------------------------- #
# build_workbook_filename
# --------------------------------------------------------------------------- #

class TestBuildWorkbookFilename(unittest.TestCase):

    def test_basic_first_and_last_name(self):
        self.assertEqual(
            ms.build_workbook_filename("Jane", "Doe"),
            "Jane_Doe_raphael_behavioral_data.xlsx",
        )

    def test_strips_unsafe_filesystem_characters(self):
        self.assertEqual(
            ms.build_workbook_filename('Mary/Jane', 'O"Brien:'),
            "MaryJane_OBrien_raphael_behavioral_data.xlsx",
        )

    def test_collapses_internal_whitespace_to_underscore(self):
        self.assertEqual(
            ms.build_workbook_filename("Mary Jane", "Van Der Berg"),
            "Mary_Jane_Van_Der_Berg_raphael_behavioral_data.xlsx",
        )

    def test_strips_leading_and_trailing_whitespace(self):
        self.assertEqual(
            ms.build_workbook_filename("  Jane  ", "  Doe  "),
            "Jane_Doe_raphael_behavioral_data.xlsx",
        )


# --------------------------------------------------------------------------- #
# _find_or_create_slot
# --------------------------------------------------------------------------- #

class TestFindOrCreateSlot(unittest.TestCase):

    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_empty_sheet_returns_row_one_new(self):
        row, is_new = ms._find_or_create_slot(self.ws, "Tantrum")
        self.assertEqual(row, 1)
        self.assertTrue(is_new)

    def test_finds_existing_match_exact_case(self):
        self.ws.cell(row=1, column=ms._VALUE_COL, value="Tantrum")
        row, is_new = ms._find_or_create_slot(self.ws, "Tantrum")
        self.assertEqual(row, 1)
        self.assertFalse(is_new)

    def test_finds_existing_match_case_insensitive_and_trimmed(self):
        self.ws.cell(row=1, column=ms._VALUE_COL, value="  Tantrum  ")
        row, is_new = ms._find_or_create_slot(self.ws, "tantrum")
        self.assertEqual(row, 1)
        self.assertFalse(is_new)

    def test_skips_filled_slots_to_next_free(self):
        self.ws.cell(row=1, column=ms._VALUE_COL, value="Tantrum")
        self.ws.cell(row=1 + ms.ROWS_PER_SLOT, column=ms._VALUE_COL, value="Social Isolation")
        row, is_new = ms._find_or_create_slot(self.ws, "Physical Aggression")
        self.assertEqual(row, 1 + 2 * ms.ROWS_PER_SLOT)
        self.assertTrue(is_new)

    def test_matches_second_occupied_slot_not_first(self):
        self.ws.cell(row=1, column=ms._VALUE_COL, value="Tantrum")
        self.ws.cell(row=1 + ms.ROWS_PER_SLOT, column=ms._VALUE_COL, value="Social Isolation")
        row, is_new = ms._find_or_create_slot(self.ws, "Social Isolation")
        self.assertEqual(row, 1 + ms.ROWS_PER_SLOT)
        self.assertFalse(is_new)


# --------------------------------------------------------------------------- #
# _chart_anchor_row
# --------------------------------------------------------------------------- #

class TestChartAnchorRow(unittest.TestCase):

    def test_string_anchor(self):
        chart = MagicMock()
        chart.anchor = "D31"
        self.assertEqual(ms._chart_anchor_row(chart), 31)

    def test_object_anchor_with_from_row(self):
        chart = MagicMock()
        chart.anchor = MagicMock()
        chart.anchor._from = MagicMock()
        chart.anchor._from.row = 60  # 0-indexed -> row 61
        self.assertEqual(ms._chart_anchor_row(chart), 61)

    def test_object_anchor_without_from_returns_none(self):
        chart = MagicMock()
        chart.anchor = MagicMock(spec=[])  # no _from attribute at all
        self.assertIsNone(ms._chart_anchor_row(chart))

    def test_malformed_string_anchor_returns_none(self):
        chart = MagicMock()
        chart.anchor = "not-a-cell-ref"
        self.assertIsNone(ms._chart_anchor_row(chart))


# --------------------------------------------------------------------------- #
# _clear_slot
# --------------------------------------------------------------------------- #

class TestClearSlot(unittest.TestCase):

    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def _populate_slot(self, anchor_row: int, category: str) -> None:
        request = ms.GraphRequest(
            graph_type=ms.GraphType.BEHAVIOR,
            category_name=category,
            data_points=make_points(3),
        )
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row)
        ms._add_chart(self.ws, request, anchor_row, first_data_row, last_data_row)

    def test_clears_values_and_chart_in_slot(self):
        self._populate_slot(1, "Tantrum")
        self.assertEqual(len(self.ws._charts), 1)

        ms._clear_slot(self.ws, 1)

        self.assertIsNone(self.ws.cell(row=1, column=ms._VALUE_COL).value)
        self.assertIsNone(self.ws.cell(row=2, column=ms._DATA_COL).value)
        self.assertEqual(len(self.ws._charts), 0)

    def test_does_not_disturb_other_slots(self):
        self._populate_slot(1, "Tantrum")
        self._populate_slot(1 + ms.ROWS_PER_SLOT, "Social Isolation")
        self.assertEqual(len(self.ws._charts), 2)

        ms._clear_slot(self.ws, 1)

        self.assertEqual(len(self.ws._charts), 1)
        self.assertEqual(
            self.ws.cell(row=1 + ms.ROWS_PER_SLOT, column=ms._VALUE_COL).value,
            "Social Isolation",
        )


# --------------------------------------------------------------------------- #
# _write_table
# --------------------------------------------------------------------------- #

class TestWriteTable(unittest.TestCase):

    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_header_and_sorted_rows(self):
        points = [ms.DataPoint(date(2026, 1, 15), 2), ms.DataPoint(date(2026, 1, 1), 9)]
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=points)
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)

        self.assertEqual(self.ws.cell(row=1, column=1).value, "Date")
        self.assertEqual(self.ws.cell(row=1, column=2).value, "Tantrum")
        # sorted by date ascending, regardless of input order
        self.assertEqual(self.ws.cell(row=2, column=1).value, date(2026, 1, 1))
        self.assertEqual(self.ws.cell(row=2, column=2).value, 9)
        self.assertEqual(self.ws.cell(row=3, column=1).value, date(2026, 1, 15))
        self.assertEqual(self.ws.cell(row=3, column=2).value, 2)
        self.assertEqual((first_data_row, last_data_row), (2, 3))

    def test_percentage_stored_as_fraction_with_percent_format(self):
        points = [ms.DataPoint(date(2026, 1, 1), 40), ms.DataPoint(date(2026, 1, 8), 72.5)]
        request = ms.GraphRequest(graph_type=ms.GraphType.REPLACEMENT, category_name="Uses Words",
                                   data_points=points)
        ms._write_table(self.ws, request, anchor_row=1)

        self.assertAlmostEqual(self.ws.cell(row=2, column=2).value, 0.40)
        self.assertAlmostEqual(self.ws.cell(row=3, column=2).value, 0.725)
        self.assertEqual(self.ws.cell(row=2, column=2).number_format, "0.0%")

    def test_zebra_banding_on_data_rows(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        ms._write_table(self.ws, request, anchor_row=1)

        # first data row (row 2) banded, second (row 3) not, third (row 4) banded
        self.assertEqual(self.ws.cell(row=2, column=1).fill.fgColor.rgb, "00DCE6F1")
        self.assertNotEqual(self.ws.cell(row=3, column=1).fill.fill_type, "solid")
        self.assertEqual(self.ws.cell(row=4, column=1).fill.fgColor.rgb, "00DCE6F1")

    def test_no_gap_band_fill_below_the_table(self):
        """
        The beige gap-band feature was removed on request -- rows below
        the table, out to the end of the 30-row slot, must be left
        unfilled rather than shaded.
        """
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)

        for r in (last_data_row + 1, 1 + ms.ROWS_PER_SLOT - 1):
            cell = self.ws.cell(row=r, column=1)
            self.assertNotEqual(cell.fill.fill_type, "solid")


# --------------------------------------------------------------------------- #
# _add_chart
# --------------------------------------------------------------------------- #

class TestAddChart(unittest.TestCase):

    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_chart_dimensions_match_spec(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        self.assertAlmostEqual(chart.height, 3.0 * 2.54, places=2)   # 3 in
        self.assertAlmostEqual(chart.width, 6.5 * 2.54, places=2)    # 6.5 in

    def test_chart_title_is_bare_category_name(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        title_text = chart.title.tx.rich.p[0].r[0].t
        self.assertEqual(title_text, "Tantrum")

    def test_percentage_chart_has_0_to_1_scaling(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.REPLACEMENT, category_name="Uses Words",
                                   data_points=make_points(3, value_fn=lambda i: 40 + i))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        self.assertEqual(chart.y_axis.scaling.min, 0)
        self.assertEqual(chart.y_axis.scaling.max, 1)

    def test_chart_anchored_at_correct_cell(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=31)
        ms._add_chart(self.ws, request, 31, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        self.assertEqual(chart.anchor, "D31")

    def test_axis_positions_are_explicit_not_openpyxl_default(self):
        """
        Regression test: openpyxl defaults BOTH x_axis.axPos and
        y_axis.axPos to "l" (left) unless set explicitly. Left
        uncorrected, Excel/LibreOffice render a garbled chart (no
        proper bottom axis, one legend entry per data point instead
        of one series). This must never regress.
        """
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        self.assertEqual(chart.x_axis.axPos, "b")
        self.assertEqual(chart.y_axis.axPos, "l")
        self.assertFalse(chart.x_axis.delete)
        self.assertFalse(chart.y_axis.delete)

    def test_chart_has_markers_enabled(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        self.assertTrue(chart.marker)
        self.assertEqual(chart.series[0].marker.symbol, "circle")

    def test_manual_layout_leaves_margin_around_plot_area(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        layout = chart.layout.manualLayout
        self.assertIsNotNone(layout)
        # plot area must not span the full 0..1 canvas -- some margin has
        # to be left on every side for the title and axis labels
        self.assertGreater(layout.x, 0)
        self.assertGreater(layout.y, 0)
        self.assertLess(layout.x + layout.w, 1.0)
        self.assertLess(layout.y + layout.h, 1.0)

    def test_no_chart_area_border(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        self.assertTrue(chart.graphical_properties.ln.noFill)

    def test_gridlines_are_thin_and_light_gray(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        line = chart.y_axis.majorGridlines.spPr.ln
        self.assertEqual(line.solidFill.srgbClr, "D9D9D9")
        self.assertLess(line.w, 12700)  # under 1pt

    def test_title_and_axis_title_are_not_bold(self):
        request = ms.GraphRequest(graph_type=ms.GraphType.BEHAVIOR, category_name="Tantrum",
                                   data_points=make_points(3))
        first_data_row, last_data_row = ms._write_table(self.ws, request, anchor_row=1)
        ms._add_chart(self.ws, request, 1, first_data_row, last_data_row)

        chart = self.ws._charts[0]
        self.assertFalse(chart.title.tx.rich.p[0].pPr.defRPr.b)
        self.assertFalse(chart.y_axis.title.tx.rich.p[0].pPr.defRPr.b)


# --------------------------------------------------------------------------- #
# generate_excel_graph — integration at the public-API level
# --------------------------------------------------------------------------- #

class TestGenerateExcelGraphIntegration(unittest.TestCase):

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp(prefix="raphael_test_")

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _req(self, gtype, cat, n_points=3, **kw):
        kw.setdefault("client_first_name", "John")
        kw.setdefault("client_last_name", "Smith")
        return ms.GraphRequest(graph_type=gtype, category_name=cat,
                                data_points=make_points(n_points), output_dir=self.tmp_dir, **kw)

    def test_creates_workbook_with_three_sheets(self):
        result = ms.generate_excel_graph(self._req(ms.GraphType.BEHAVIOR, "Tantrum"))
        wb = load_workbook(result.workbook_path)
        self.assertEqual(set(wb.sheetnames), {"Behavior", "Replacement", "Caregiver's Goal"})

    def test_new_category_lands_at_row_one(self):
        result = ms.generate_excel_graph(self._req(ms.GraphType.BEHAVIOR, "Tantrum"))
        self.assertEqual(result.anchor_row, 1)
        self.assertFalse(result.was_replacement)

    def test_second_category_lands_at_next_slot(self):
        ms.generate_excel_graph(self._req(ms.GraphType.BEHAVIOR, "Tantrum"))
        result = ms.generate_excel_graph(self._req(ms.GraphType.BEHAVIOR, "Social Isolation"))
        self.assertEqual(result.anchor_row, 1 + ms.ROWS_PER_SLOT)

    def test_overwrite_same_category_replaces_in_place(self):
        first = ms.generate_excel_graph(self._req(ms.GraphType.BEHAVIOR, "Tantrum", n_points=3))
        second = ms.generate_excel_graph(
            self._req(ms.GraphType.BEHAVIOR, "tantrum", n_points=2)  # different case, fewer rows
        )
        self.assertEqual(first.anchor_row, second.anchor_row)
        self.assertTrue(second.was_replacement)

        wb = load_workbook(second.workbook_path)
        ws = wb["Behavior"]
        # only 2 data rows should remain -- row 4 (which had data after the
        # first, 3-row write) must have been cleared by the overwrite
        self.assertIsNotNone(ws.cell(row=3, column=1).value)
        self.assertIsNone(ws.cell(row=4, column=1).value)
        self.assertEqual(len(ws._charts), 1)

    def test_persists_and_reloads_across_separate_calls(self):
        r1 = ms.generate_excel_graph(self._req(ms.GraphType.BEHAVIOR, "Tantrum"))
        r2 = ms.generate_excel_graph(self._req(ms.GraphType.REPLACEMENT, "Uses Words"))
        self.assertEqual(r1.workbook_path, r2.workbook_path)

        wb = load_workbook(r2.workbook_path)
        self.assertEqual(wb["Behavior"].cell(row=1, column=2).value, "Tantrum")
        self.assertEqual(wb["Replacement"].cell(row=1, column=2).value, "Uses Words")

    def test_too_many_points_raises_before_writing(self):
        req = self._req(ms.GraphType.BEHAVIOR, "Overflow",
                         n_points=ms.MAX_DATA_POINTS_PER_TABLE + 1)
        expected_path = os.path.join(self.tmp_dir, ms.build_workbook_filename("John", "Smith"))
        with self.assertRaises(ms.GraphRequestError):
            ms.generate_excel_graph(req)
        self.assertFalse(os.path.exists(expected_path))

    def test_explicit_workbook_path_overrides_default(self):
        custom_path = os.path.join(self.tmp_dir, "custom_name.xlsx")
        result = ms.generate_excel_graph(
            self._req(ms.GraphType.BEHAVIOR, "Tantrum", workbook_path=custom_path)
        )
        self.assertEqual(result.workbook_path, custom_path)
        self.assertTrue(os.path.exists(custom_path))

    def test_default_filename_uses_client_first_and_last_name(self):
        result = ms.generate_excel_graph(
            self._req(ms.GraphType.BEHAVIOR, "Tantrum",
                      client_first_name="Jane", client_last_name="Doe")
        )
        self.assertEqual(os.path.basename(result.workbook_path),
                          "Jane_Doe_raphael_behavioral_data.xlsx")

    def test_filename_sanitizes_unsafe_characters_and_whitespace(self):
        result = ms.generate_excel_graph(
            self._req(ms.GraphType.BEHAVIOR, "Tantrum",
                      client_first_name="Mary Jane", client_last_name='O"Brien:')
        )
        self.assertEqual(os.path.basename(result.workbook_path),
                          "Mary_Jane_OBrien_raphael_behavioral_data.xlsx")

    def test_different_clients_get_different_workbooks(self):
        r1 = ms.generate_excel_graph(
            self._req(ms.GraphType.BEHAVIOR, "Tantrum",
                      client_first_name="Jane", client_last_name="Doe")
        )
        r2 = ms.generate_excel_graph(
            self._req(ms.GraphType.BEHAVIOR, "Tantrum",
                      client_first_name="John", client_last_name="Smith")
        )
        self.assertNotEqual(r1.workbook_path, r2.workbook_path)
        self.assertTrue(os.path.exists(r1.workbook_path))
        self.assertTrue(os.path.exists(r2.workbook_path))


# --------------------------------------------------------------------------- #
# open_file
# --------------------------------------------------------------------------- #

class TestOpenFile(unittest.TestCase):

    @patch("main_script.subprocess.run")
    def test_uses_open_on_darwin(self, mock_run):
        with patch.object(ms.sys, "platform", "darwin"):
            ms.open_file("/tmp/some.xlsx")
        mock_run.assert_called_once_with(["open", "/tmp/some.xlsx"], check=False)

    @patch("main_script.subprocess.run")
    def test_uses_xdg_open_on_linux(self, mock_run):
        with patch.object(ms.sys, "platform", "linux"), patch.object(ms.os, "name", "posix"):
            ms.open_file("/tmp/some.xlsx")
        mock_run.assert_called_once_with(["xdg-open", "/tmp/some.xlsx"], check=False)

    def test_uses_startfile_on_windows(self):
        with patch.object(ms.sys, "platform", "win32"), patch.object(ms.os, "name", "nt"):
            ms.os.startfile = MagicMock()  # not present on non-Windows, so add it
            ms.open_file("/tmp/some.xlsx")
            ms.os.startfile.assert_called_once_with("/tmp/some.xlsx")


# --------------------------------------------------------------------------- #
# Full backend report — many tables, many rows, all three graph types
# --------------------------------------------------------------------------- #

class TestFullBackendReport(unittest.TestCase):
    """
    Drives generate_excel_graph() the way the rest of Raphael (or any
    external caller) eventually will: repeated calls, several
    categories per sheet, several rows per category, and one
    overwrite -- all landing in a single reviewable workbook at
    tests_output/excel_graph_creation_tests.xlsx.
    """

    @classmethod
    def setUpClass(cls):
        os.makedirs(TESTS_OUTPUT_DIR, exist_ok=True)
        cls.workbook_path = os.path.join(TESTS_OUTPUT_DIR, FULL_BACKEND_FILENAME)
        if os.path.exists(cls.workbook_path):
            os.remove(cls.workbook_path)

    def _generate(self, gtype, cat, points, **kw):
        req = ms.GraphRequest(graph_type=gtype, category_name=cat, data_points=points,
                               workbook_path=self.workbook_path, **kw)
        return ms.generate_excel_graph(req)

    def test_full_backend_multi_table_report(self):
        behavior_tables = {
            "Tantrum": make_points(5, value_fn=lambda i: [3, 5, 2, 4, 1][i]),
            "Social Isolation": make_points(3, value_fn=lambda i: [1, 2, 0][i]),
            "Physical Aggression": make_points(7, value_fn=lambda i: [4, 6, 5, 3, 2, 4, 3][i]),
        }
        replacement_tables = {
            "Uses Words to Request Break": make_points(4, value_fn=lambda i: [40, 55, 60, 72][i]),
            "Raises Hand": make_points(6, value_fn=lambda i: [20, 25, 30, 45, 50, 55][i]),
        }
        goal_tables = {
            "Consistent Redirection": make_points(3, value_fn=lambda i: [20, 35, 50][i]),
            "Uses Visual Schedule": make_points(4, value_fn=lambda i: [10, 30, 45, 60][i]),
        }

        results = {}
        for cat, points in behavior_tables.items():
            results[("Behavior", cat)] = self._generate(ms.GraphType.BEHAVIOR, cat, points)
        for cat, points in replacement_tables.items():
            results[("Replacement", cat)] = self._generate(ms.GraphType.REPLACEMENT, cat, points)
        for cat, points in goal_tables.items():
            results[("Caregiver's Goal", cat)] = self._generate(ms.GraphType.CAREGIVER_GOAL, cat, points)

        # Overwrite one existing category with new, shorter data -- this must
        # replace the row-1 Tantrum table in place, not create a 4th slot.
        overwrite_result = self._generate(
            ms.GraphType.BEHAVIOR, "tantrum",
            make_points(2, start=date(2026, 3, 1), value_fn=lambda i: [9, 1][i]),
        )
        self.assertEqual(overwrite_result.anchor_row, results[("Behavior", "Tantrum")].anchor_row)
        self.assertTrue(overwrite_result.was_replacement)

        # --- Verify the saved file structurally -----------------------------
        wb = load_workbook(self.workbook_path)
        self.assertEqual(set(wb.sheetnames), {"Behavior", "Replacement", "Caregiver's Goal"})

        ws_behavior = wb["Behavior"]
        self.assertEqual(len(ws_behavior._charts), 3)  # overwrite must not add a 4th
        expected_behavior_rows = {"Tantrum": 1, "Social Isolation": 31, "Physical Aggression": 61}
        for cat, expected_row in expected_behavior_rows.items():
            # Tantrum was overwritten with lowercase "tantrum" below, so compare
            # case-insensitively -- the slot-matching logic is case-insensitive
            # by design, but the literal text written is whatever was last sent.
            actual = ws_behavior.cell(row=expected_row, column=2).value
            self.assertEqual(actual.strip().lower(), cat.lower())

        # Overwritten Tantrum table should now have exactly 2 data rows.
        self.assertIsNotNone(ws_behavior.cell(row=2, column=1).value)
        self.assertIsNotNone(ws_behavior.cell(row=3, column=1).value)
        self.assertIsNone(ws_behavior.cell(row=4, column=1).value)

        ws_replacement = wb["Replacement"]
        self.assertEqual(len(ws_replacement._charts), 2)
        self.assertEqual(ws_replacement.cell(row=1, column=2).value, "Uses Words to Request Break")
        self.assertEqual(ws_replacement.cell(row=31, column=2).value, "Raises Hand")
        # spot-check a percentage value was stored as a fraction
        self.assertAlmostEqual(ws_replacement.cell(row=2, column=2).value, 0.40)

        ws_goals = wb["Caregiver's Goal"]
        self.assertEqual(len(ws_goals._charts), 2)
        self.assertEqual(ws_goals.cell(row=1, column=2).value, "Consistent Redirection")
        self.assertEqual(ws_goals.cell(row=31, column=2).value, "Uses Visual Schedule")

        # File is intentionally left on disk at self.workbook_path for
        # manual review -- not cleaned up here.
        print(f"\nFull backend report written to: {self.workbook_path}")


if __name__ == "__main__":
    unittest.main(verbosity=2)